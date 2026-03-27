import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def _get_temp_comment(path: str = "stDataPack.txt") -> str:
    """Return comment template for temperature sensors from stDataPack.txt."""
    try:
        with open(path, encoding="utf-8") as f:
            for line in f:
                if "Датчик температуры" in line and "//" in line:
                    return line.split("//", 1)[-1].strip()
    except FileNotFoundError:
        pass
    return "Датчик температуры {num}"


def _format_temp_comment(template: str, index: int, start_index: int = 0) -> str:
    """Format comment according to placeholders used in map generation."""

    def repl(match: re.Match) -> str:
        inner = match.group(1)
        if inner == "num":
            return str(index)
        parts = [p.strip() for p in inner.split(',')]
        idx = index - start_index
        if 0 <= idx < len(parts):
            return parts[idx]
        return match.group(0)

    return re.sub(r"\{([^}]+)\}", repl, template)

def generate_protocol_excel(mapping: dict, params: dict):
    """Генерирует Excel-файл protocol.xlsx с листами DI_Prot, DO_Prot, AI_Prot."""
    try:
        # Параметры
        typeQFD = params['typeQFD']
        key_auto = params['key_auto']
        num_qf = params['num_qf']
        num_fd = params['num_fd']
        num_qfd = params['num_qfd']
        num_cont_auto = params['num_cont_auto']
        num_cont_manu = params['num_cont_manu']
        num_km = params['num_km']
        count_Dt = params['count_Dt']

        ai_mapping = mapping['ai_mapping']
        di_mapping = mapping['di_mapping']
        do_mapping = mapping['do_mapping']
        lines_on_idx = mapping['lines_on_index']
        al_all_idx = mapping['al_all_index']
        al_dt_idx = mapping['al_dt_index']
        manuon_idx = mapping['manuon_index']
        auto_idx = mapping['auto_index']
        rkn_idx = mapping['rkn_index']

        # Протокол AI (аналоговые входы)
        ai_prot = [['№', 'Параметр', 'Модуль', 'Канал', 'Клеммник', 'Контакт', 'Цепь',
                    'Соответствие на мнемосхеме', 'Результат проверки канала']]
        num = 0
        # Записи по каждому сигналу AI (токи фаз и утечки)
        for klemm, line_num, phase_code, mod_num, ch_num in ai_mapping:
            num += 1
            if klemm.startswith('ta') and not klemm.startswith('tan'):
                # Определяем фазу буквенно
                if phase_code == 1:
                    phase_letter = 'A'
                elif phase_code == 2:
                    phase_letter = 'B'
                elif phase_code == 3:
                    phase_letter = 'C'
                else:
                    phase_letter = ''
                param_name = f"Ток линии {typeQFD.upper()}{line_num} фаза {phase_letter}"
                buf = [num, param_name, f"A3.{mod_num}", ch_num, klemm.upper(), 1, '24В,4..20мА', 'соотв.', 'исправен']
            elif klemm.startswith('tan'):
                param_name = f"Ток утечки линии {typeQFD.upper()}{line_num}"
                buf = [num, param_name, f"A3.{mod_num}", ch_num, klemm.upper(), 1, '24В,4..20мА', 'соотв.', 'исправен']
            else:
                # Сигналы, которые не распознаны как токи (не должно происходить обычно)
                param_name = klemm.upper()
                buf = [num, param_name, f"A3.{mod_num}", ch_num, klemm.upper(), '', '', 'соотв.', 'исправен']
            ai_prot.append(buf)
        # Добавляем протокол датчиков температуры (RTD)
        comment_template = _get_temp_comment()
        num_dt = 0
        num_mod = 1
        num_ch = 4
        flag = False
        for i in range(count_Dt + 1):
            num += 1
            desc = _format_temp_comment(comment_template, num_dt)
            buf = [num, desc, f"A4.{num_mod}", num_ch,
                   "XT3", f"{num_dt*4+1}...{num_dt*4+4}", "термосопротивление", "соотв.", "исправен"]
            ai_prot.append(buf)
            num_dt += 1
            num_ch += 1
            if num_ch == 4 and flag is True:
                num_mod += 1
            if num_ch == 5:
                num_ch = 1
                flag = True

        # Протокол DI (цифровые входы)
        di_prot = [['№', 'Параметр', 'Модуль', 'Канал', 'Реле', 'Контакт',
                    'Цепь', 'Соответствие на мнемосхеме', 'Результат проверки канала']]
        num = 0
        # Основные DI-сигналы (QF, QFD, FD, KM)
        for sig_name, sig_type, number, mod_num, ch_num in di_mapping:
            num += 1
            # Определяем контакт (номер) в зависимости от типа сигнала
            if sig_type == 'qfd':
                contact = number
            elif sig_type == 'qf':
                contact = num_qf
            elif sig_type == 'fd':
                contact = num_fd
            elif sig_type == 'km':
                contact = num_km
            else:
                contact = '---'
            # Параметр (описание)
            if sig_type in ('qf', 'fd', 'qfd'):
                param_name = f"Срабатывание {sig_name.upper()}"
            elif sig_type == 'km':
                param_name = f"Вкл/откл линии №{sig_name[2:]}"
            else:
                param_name = sig_name.upper()
            di_prot.append([num, param_name, f"A1.{mod_num}", ch_num + 1, sig_name.upper(),
                            contact, '24В, вх.', 'соответств.', 'исправен'])
        # Добавляем сигнал "Обогрев в ручн. режиме"
        num += 1
        di_prot.append([num, 'Обогрев в ручн. режиме', f"A1.{manuon_idx[0]}", manuon_idx[1] + 1,
                        key_auto, num_cont_manu, '24В, вх.', 'соответств.', 'исправен'])
        # Сигнал "Обогрев в авт. режиме"
        num += 1
        di_prot.append([num, 'Обогрев в авт. режиме', f"A1.{auto_idx[0]}", auto_idx[1] + 1,
                        key_auto, num_cont_auto, '24В, вх.', 'соответств.', 'исправен'])
        # Сигнал РКН (наличие напряжения на вводе), если есть
        num += 1
        if rkn_idx is None:
            # Если РКН отсутствует, пропускаем добавление (номер просто не используется)
            # Можно при необходимости залогировать предупреждение
            pass
        else:
            di_prot.append([num, 'Наличие напряжения на вводе', f"A1.{rkn_idx[0]}",
                            rkn_idx[1] + 1, 'XXXX', 'XXXX', '24В, вх.', 'соответств.', 'исправен'])

        # Протокол DO (цифровые выходы)
        do_prot = [['№', 'Параметр', 'Модуль', 'Канал', 'Цепь',
                    'Соответствие на мнемосхеме', 'Результат проверки канала']]
        num = 0
        num_mod = 1
        num_ch = 1
        # Основные DO-сигналы (KM выходы)
        for entry in do_mapping:
            num += 1
            # Формируем параметр: "Вкл/откл линии №X" (номер линии или с подномер)
            sig_name = entry[0]  # 'kmX' или 'kmX.Y'
            param_name = f"Вкл/откл линии №{sig_name[2:]}"
            do_prot.append([num, param_name, f"A2.{num_mod}", num_ch, '24В, вх.', 'соответств.', 'исправен'])
            num_ch += 1
            if num_ch > 16:
                num_ch = 0
                num_mod += 1
        # Специальные DO-сигналы (LinesOn, AlarmAll, AlarmDt)
        num += 1
        do_prot.append([num, 'Обогрев включен', f"A2.{lines_on_idx[0]}", lines_on_idx[1],
                        '24В, вх.', 'соответств.', 'исправен'])
        num += 1
        do_prot.append([num, 'Общая авария', f"A2.{al_all_idx[0]}", al_all_idx[1],
                        '24В, вх.', 'соответств.', 'исправен'])
        num += 1
        do_prot.append([num, 'Общая авария датчика', f"A2.{al_dt_idx[0]}", al_dt_idx[1],
                        '24В, вх.', 'соответств.', 'исправен'])

        # Создаем DataFrame для каждого протокола
        df_di_prot = pd.DataFrame(di_prot[1:], columns=di_prot[0])
        df_do_prot = pd.DataFrame(do_prot[1:], columns=do_prot[0])
        df_ai_prot = pd.DataFrame(ai_prot[1:], columns=ai_prot[0])
        # Записываем в Excel-файл с использованием XlsxWriter
        sheets = ['DI_Prot', 'DO_Prot', 'AI_Prot']
        excel_path = "protocol.xlsx"
        with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
            df_di_prot.to_excel(writer, sheet_name='DI_Prot', index=False)
            df_do_prot.to_excel(writer, sheet_name='DO_Prot', index=False)
            df_ai_prot.to_excel(writer, sheet_name='AI_Prot', index=False)
            # Форматирование ширины столбцов и переноса текста
            workbook = writer.book
            cell_format = workbook.add_format({
                'text_wrap': True,
                'font_name': 'Times New Roman',
                'font_size': 11
            })
            for sheet_name in sheets:
                worksheet = writer.sheets[sheet_name]
                worksheet.set_column('A:A', 3.71, cell_format)
                worksheet.set_column('B:B', 39.57, cell_format)
                worksheet.set_column('C:C', 9.71, cell_format)
                worksheet.set_column('D:D', 12.71, cell_format)
                worksheet.set_column('E:E', 9.14, cell_format)
                worksheet.set_column('F:F', 12.71, cell_format)
                worksheet.set_column('G:G', 12.71, cell_format)
                worksheet.set_column('H:H', 16.14, cell_format)
                worksheet.set_column('I:I', 16.14, cell_format)
                worksheet.set_column('J:J', 12.57, cell_format)
        # Открываем с помощью openpyxl для применения выравнивания текста
        wb = load_workbook(excel_path)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = alignment
        wb.save(excel_path)
    except Exception as e:
        raise Exception(f"Ошибка формирования protocol.xlsx: {e}")
